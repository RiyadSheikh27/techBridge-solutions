import openpyxl
import uuid
from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets
from .models import *
from .serializers import *
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny, IsAuthenticatedOrReadOnly
from rest_framework.decorators import action
from openpyxl import load_workbook
from io import BytesIO
from django.core.cache import cache
from decimal import Decimal, InvalidOperation
from django.db import transaction
from authentication.permissions import IsOwnerOrReadOnly
from checkout.models import Order, ProductReview

""" Start of Creating Views for Product Section """

class CustomResponseMixin:
    """Custom response mixin for API responses"""
    def success_response(self, data=None, message="Success", status_code=status.HTTP_200_OK):
        return Response({
            "success": True,
            "message": message,
            "data": data
        }, status=status_code)

    def error_response(self, message="Error occurred", errors=None, status_code=status.HTTP_400_BAD_REQUEST):
        response_data = {
            'success': False,
            'message': message,
            'data': None
        }
        if errors:
            response_data['errors'] = errors
        return Response(response_data, status=status_code)

class ProductCategoryViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Product Categories
    
    list: Get all categories with subcategories and products
    retrieve: Get a single category with all details
    create: Create a new category
    update: Update a category
    partial_update: Partially update a category
    destroy: Delete a category
    """
    queryset = ProductCategory.objects.filter(is_active=True)
    permission_classes = [IsAuthenticatedOrReadOnly]
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductCategoryWriteSerializer
        return ProductCategorySerializer
    
    def get_queryset(self):
        queryset = ProductCategory.objects.all()
        
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        product_type = self.request.query_params.get('product_type', None)
        if product_type:
            queryset = queryset.filter(product_type=product_type)
        
        return queryset.order_by('display_order', 'name')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Category Listed Successfully"
            )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return self.success_response(
            data=serializer.data,
            message="Category retrieved successfully"
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data= request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Category created successfully",
                status_code = status.HTTP_201_CREATED
            )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Category updated successfully"
            )
        return self.error_response(
            message="Failed to update category",
            errors=serializere.rrors
        )
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.success_response(
            message="Category deleted successfully"
        )

class ProductSubCategoryViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Product SubCategories
    
    list: Get all subcategories with products
    retrieve: Get a single subcategory with all details
    create: Create a new subcategory
    update: Update a subcategory
    partial_update: Partially update a subcategory
    destroy: Delete a subcategory
    by_category: Get subcategories filtered by category slug
    """
    queryset = ProductSubCategory.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductSubCategoryWriteSerializer
        return ProductSubCategorySerializer

    def get_queryset(self):
        queryset = ProductSubCategory.objects.all()
        
        """ Filter by category """
        category_slug = self.request.query_params.get('category', None)
        if category_slug:
            queryset = queryset.filter(category__slug=category_slug)
        
        """ Filter by active status """
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('display_order', 'name')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Subcategories retrieved successfully"
        )
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return self.success_response(
            data=serializer.data,
            message="Subcategory retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Subcategory created successfully",
                status_code=status.HTTP_201_CREATED
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Subcategory updated successfully"
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.success_response(
            message="Subcategory deleted successfully",
            status_code=status.HTTP_204_NO_CONTENT
        )

class ProductViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Products
    
    list: Get all products
    retrieve: Get a single product with full details
    create: Create a new product
    update: Update a product
    partial_update: Partially update a product
    destroy: Delete a product
    by_type: Get products grouped by product type
    by_subcategory: Get products by subcategory slug
    featured: Get featured products
    """
    queryset = Product.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductWriteSerializer
        elif self.action == 'list':
            return ProductListSerializer
        return ProductDetailSerializer
        

    def get_queryset(self):
        queryset = Product.objects.select_related(
            'subcategory',
            'subcategory__category'
        ).prefetch_related(
            'productdescription_set',
            'productdescription_set__productdescriptionrow_set'
        )
        
        """Filter by subcategory"""
        subcategory_slug = self.request.query_params.get('subcategory', None)
        if subcategory_slug:
            queryset = queryset.filter(subcategory__slug=subcategory_slug)
        
        """Filter by category"""
        category_slug = self.request.query_params.get('category', None)
        if category_slug:
            queryset = queryset.filter(subcategory__category__slug=category_slug)
        
        """Filter by active status"""
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        """Filter by featured"""
        is_featured = self.request.query_params.get('is_featured', None)
        if is_featured is not None:
            queryset = queryset.filter(is_featured=is_featured.lower() == 'true')
        
        """Search by name or manufacturer"""
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(manufacturer__icontains=search) |
                Q(mfr_part__icontains=search)
            )
        
        return queryset.order_by('display_order', '-created_at')
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Products retrieved successfully"
        )
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return self.success_response(
            data=serializer.data,
            message="Product retrieved successfully"
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Product created successfully",
                status_code=status.HTTP_201_CREATED
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Product updated successfully"
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.success_response(
            message="Product deleted successfully",
            status_code=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'])
    def by_subcategory(self, request):
        """Get products by subcategory slug"""
        subcategory_slug = request.query_params.get('slug')
        if not subcategory_slug:
            return self.error_response(
                message="Subcategory slug is required"
            )
        
        products = self.get_queryset().filter(
            subcategory__slug=subcategory_slug
        )
        serializer = ProductDetailSerializer(products, many=True)
        return self.success_response(
            data=serializer.data,
            message="Products retrieved successfully"
        )
    
    @action(detail=False, methods=['get'])
    def featured(self, request):
        """Get featured products"""
        products = self.get_queryset().filter(is_featured=True)
        serializer = self.get_serializer(products, many=True)
        return self.success_response(
            data=serializer.data,
            message="Featured products retrieved successfully"
        )

class CategoryDescriptionViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Category Descriptions

    list: Get all category descriptions
    retrieve: Get a single category description
    create: Create a new category description
    update: Update a category description
    partial_update: Partially update a category description
    destroy: Delete a category description
    """
    queryset = CategoryDescription.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return CategoryDescriptionWriteSerializer
        return CategoryDescriptionSerializer

    def get_queryset(self):
        queryset = CategoryDescription.objects.all()
    
        subcategory_slug = self.request.query_params.get('subcategory', None)
        if subcategory_slug:
            queryset = queryset.filter(productSubCategory__slug=subcategory_slug)
    
        return queryset
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Category descriptions retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Category description created successfully",
                status_code=status.HTTP_201_CREATED
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Category description updated successfully"
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Category description partially updated successfully"
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.success_response(
            message="Category description deleted successfully",
            status_code=204
        )

class ProductDescriptionViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Product Descriptions

    list: Get all product descriptions
    retrieve: Get a single product description
    create: Create a new product description
    update: Update a product description
    partial_update: Partially update a product description
    destroy: Delete a product description
    """
    queryset = ProductDescription.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductDescriptionWriteSerializer
        return ProductDescriptionSerializer
    
    def get_queryset(self):
        queryset = ProductDescription.objects.all()
        
        """Filter by product"""
        product_slug = self.request.query_params.get('product', None)
        if product_slug:
            queryset = queryset.filter(product__slug=product_slug)
        
        return queryset.order_by('display_order')
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Product descriptions retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Product description created successfully",
                status_code=status.HTTP_201_CREATED
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data = request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data = serializer.data,
                message = "Product Description Updated Successfully"
            )
        return self.error_response(
            message = "Validation failed",
            errors = serializer.errors
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data = serializer.data,
                message = "Product Description Updated Successfully"
            )
        return self.error_response(
            message = "Validation failed",
            errors = serializer.errors
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()

        return self.success_response(
            message="Category Description Deleted Sucessfully",
            status_code=204
        )



class ProductDescriptionRowViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet for Product Description Rows

    list: Get all product description rows
    retrieve: Get a single product description row
    create: Create a new product description row
    update: Update a product description row
    partial_update: Partially update a product description row
    destroy: Delete a product description row
    """
    queryset = ProductDescriptionRow.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductDescriptionRowWriteSerializer
        return ProductDescriptionRowSerializer
    
    def get_queryset(self):
        queryset = ProductDescriptionRow.objects.all()
        
        """Filter by description"""
        description_id = self.request.query_params.get('description', None)
        if description_id:
            queryset = queryset.filter(description__id=description_id)
        
        return queryset.order_by('display_order')
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Product description rows retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Product description row created successfully",
                status_code=status.HTTP_201_CREATED
            )
        return self.error_response(
            message="Validation failed",
            errors=serializer.errors
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data = serializer.data,
                message="Product Description Updated Successfully"
            )

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data = serializer.data,
                message="Product Description Updated Successfully"
            )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return self.success_response(
            status_code=204,
            message="Product Description Row Deleted Successfully"
        )

class ProductReviewViewSet(CustomResponseMixin, viewsets.ModelViewSet):
    """ViewSet for Product Reviews"""
    permission_classes = [IsAuthenticatedOrReadOnly, IsOwnerOrReadOnly]
    lookup_field = 'id'

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProductReviewWriteSerializer
        return ProductReviewSerializer

    def get_queryset(self):
        queryset = ProductReview.objects.select_related('order', 'user')[:10]
        
        """ Filter by Order"""
        order_id = self.request.query_params.get('order')
        if order_id:
            get_object_or_404(Order, id=order_id)
            queryset = queryset.filter(order_id=order_id)
        
        """ Filter by user for their own reviews"""
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
            
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(
            data=serializer.data,
            message="Product reviews retrieved successf ully"
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            """Auto-assign the authenticated user"""
            serializer.save(user=request.user)
            return self.success_response(
                data=serializer.data,
                message="Product Review Created Successfully",
                status_code=201
            )
        return self.error_response(
            message="Validation Failed",
            errors=serializer.errors
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        
        """ Check if user owns this review"""
        if instance.user != request.user:
            return self.error_response(
                message="You can only update your own reviews",
                status_code=403
            )
        
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return self.success_response(
                data=serializer.data,
                message="Product Review Updated Successfully"
            )
        return self.error_response(
            message="Validation Failed",
            errors=serializer.errors
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        
        """ Check if user owns this review"""
        if instance.user != request.user and request.user.role != 'admin':
            return self.error_response(
                message="You can only delete your own reviews",
                status_code=403
            )
        
        instance.delete()
        return self.success_response(
            status_code=204,
            message="Product Review Deleted Successfully"
        )

""" Bulk Product Upload from Excel Section """
class BulkProductUploadViewSet(CustomResponseMixin, viewsets.ViewSet):
    """ViewSet for Bulk Product Upload from Excel"""
    permission_classes = [IsAdminUser]
    
    def _safe_decimal(self, value, default='0.00'):
        """Safe Decimal Conversion"""
        if value is None or value == '':
            return Decimal(default)
        try:
            return Decimal(str(value))
        except (ValueError, InvalidOperation):
            return Decimal(default)

    def _safe_int(self, value, default=0):
        """Safely convert value to integer"""
        if value is None or value == '':
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_bool(self, value, default=True):
        """Safely convert value to boolean"""
        if value is None or value == '':
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['true', 'yes', '1', 'y']
        return bool(value)

    def _parse_excel_row(self, row, subcategory_id, row_number):
        """Parse a single row from the Excel file"""
        try:
            temp_id = str(uuid.uuid4())

            product_data = {
                'id': temp_id,
                'row_number': row_number,
                'subcategory': subcategory_id,
                'name': str(row[0].value or '').strip(),  # Column A
                'series': str(row[1].value or '').strip() if row[1].value else None,  # Column B
                'msrp': str(self._safe_decimal(row[2].value, '0.00')),  # Column C
                'price': str(self._safe_decimal(row[3].value, '0.00')),  # Column D
                'stock': self._safe_int(row[4].value, 100),  # Column E
                'is_in_stock': self._safe_bool(row[5].value, True),  # Column F
                'mfr_part': str(row[6].value or '').strip() if row[6].value else None,  # Column G
                'shi_part': str(row[7].value or '').strip() if row[7].value else None,  # Column H
                'unspsc': str(row[8].value or '').strip() if row[8].value else None,  # Column I
                'manufacturer': str(row[9].value or '').strip() if row[9].value else None,  # Column J
                'description': str(row[10].value or '').strip(),  # Column K
                'is_active': self._safe_bool(row[11].value, False),  # Column L
                'is_featured': self._safe_bool(row[12].value, False),  # Column M
                'display_order': self._safe_int(row[13].value, 0),  # Column N
                'image': str(row[14].value or '').strip() if row[14].value else None,  # Column O
                'valid': True,
                'errors': []
            }
            
            """ Validate required fields"""
            if not product_data['name']:
                product_data['valid'] = False
                product_data['errors'].append('Product Name is required')

            if Decimal(product_data['msrp']) <= 0:
                product_data['valid'] = False
                product_data['errors'].append('MSRP must be greater than 0')

            if Decimal(product_data['price']) <= 0:
                product_data['valid'] = False
                product_data['errors'].append('Product Price must be greater than 0')

            if not product_data['description']:
                product_data['valid'] = False
                product_data['errors'].append('Product Description is required')

            return product_data

        except Exception as e:
            return {
                'id': str(uuid.uuid4()),
                'row_number': row_number,
                'valid': False,
                'errors': [f'Error parsing row: {str(e)}'],
                'name': 'Error'
            }
    
    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """
        Upload Excel file and get parsed products with unique IDs
        Expected request:
        - subcategory (UUID): Subcategory ID
        - file (File): Excel file
        """
        subcategory_id = request.data.get('subcategory')
        excel_file = request.data.get('file')

        if not subcategory_id:
            return self.error_response(
                message="Subcategory ID is required"
            )
        
        if not excel_file:
            return self.error_response(
                message="Excel file is required"
            )
        """Validate Subcategory exists"""
        try:
            subcategory = ProductSubCategory.objects.get(id=subcategory_id)
        except ProductSubCategory.DoesNotExist:
            return self.error_response(
                message="Subcategory does not exist",
                status_code=status.HTTP_404_NOT_FOUND
            )
        try:
            """Read Excel file"""
            workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))
            sheet = workbook.active
            
            """Parse products (skip header row)"""
            products = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                """Skip empty rows"""
                if not any(cell.value for cell in row):
                    continue
                
                product_data = self._parse_excel_row(row, subcategory_id, idx)
                products.append(product_data)
            
            if not products:
                return self.error_response(
                    message="No valid products found in Excel file"
                )

            """ Store Products in Cache for Later Retrieval (30 minutes)"""
            cache_key = f"bulk_upload_{request.user.id}_{uuid.uuid4()}"
            cache.set(cache_key, products, timeout=1800)

            """Summary"""
            valid_count = sum(1 for p in products if p['valid'])
            invalid_count = len(products) - valid_count

            return self.success_response(
                data={
                    'cache_key': cache_key,
                    'subcategory': {
                        'id': str(subcategory.id),
                        'name': subcategory.name,
                        'category_name': subcategory.category.name
                    },
                    'products': products,
                    'summary': {
                        'total': len(products),
                        'valid': valid_count,
                        'invalid': invalid_count
                    }
                },
                message=f"Excel parsed successfully. {valid_count} valid products found."
            )
            
        except openpyxl.utils.exceptions.InvalidFileException:
            return self.error_response(
                message="Invalid Excel file format. Please upload a valid .xlsx file"
            )
        except Exception as e:
            return self.error_response(
                message=f"Failed to parse Excel file: {str(e)}"
            )
    
    @action(detail=False, methods=['post'])
    def save_selected(self, request):
        """
        Step 2: Save selected products to database
        
        Expected request:
        - cache_key (string): Cache key from upload_excel response
        - selected_ids (List): List of product IDs to create
        """
        cache_key = request.data.get('cache_key')
        selected_ids = request.data.get('selected_ids', [])
        
        if not cache_key:
            return self.error_response(
                message="Cache key is required"
            )
        
        if not selected_ids or not isinstance(selected_ids, list):
            return self.error_response(
                message="Selected IDs must be a non-empty list"
            )
        
        """Retrieve products from cache"""
        cached_products = cache.get(cache_key)
        if not cached_products:
            return self.error_response(
                message="Upload session expired or not found. Please upload Excel again.",
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        """Filter selected products"""
        products_to_create = [
            p for p in cached_products 
            if p['id'] in selected_ids and p['valid']
        ]
        
        if not products_to_create:
            return self.error_response(
                message="No valid products selected"
            )
        
        created_products = []
        errors = []
        
        try:
            with transaction.atomic():
                for product_data in products_to_create:
                    try:
                        """Remove temporary fields"""
                        temp_id = product_data['id']
                        product_data.pop('id', None)
                        product_data.pop('row_number', None)
                        product_data.pop('valid', None)
                        product_data.pop('errors', None)
                        
                        """Convert string decimals to Decimal"""
                        product_data['msrp'] = Decimal(product_data['msrp'])
                        product_data['price'] = Decimal(product_data['price'])
                        
                        """Get subcategory"""
                        subcategory = ProductSubCategory.objects.get(
                            id=product_data['subcategory']
                        )
                        product_data['subcategory'] = subcategory
                        
                        """Create product"""
                        product = Product.objects.create(**product_data)
                        created_products.append(product)
                        
                    except ProductSubCategory.DoesNotExist:
                        errors.append({
                            'temp_id': temp_id,
                            'name': product_data.get('name', 'Unknown'),
                            'error': 'Subcategory not found'
                        })
                    except Exception as e:
                        errors.append({
                            'temp_id': temp_id,
                            'name': product_data.get('name', 'Unknown'),
                            'error': str(e)
                        })
                
                """If any errors occurred, rollback transaction"""
                if errors:
                    raise Exception("Some products failed to create")
            
            """Clear cache after successful creation"""
            cache.delete(cache_key)
            
            """Serialize created products"""
            serializer = ProductDetailSerializer(created_products, many=True)
            
            return self.success_response(
                data={
                    'created_products': serializer.data,
                    'summary': {
                        'total_selected': len(selected_ids),
                        'successfully_created': len(created_products),
                        'failed': len(errors)
                    },
                    'errors': errors if errors else None
                },
                message=f"Successfully created {len(created_products)} products",
                status_code=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            return self.error_response(
                message=f"Failed to create products: {str(e)}",
                errors=errors if errors else None
            )

""" End ofBulk Product Upload from Excel Section """

""" End of Creating Views for Product Section """
